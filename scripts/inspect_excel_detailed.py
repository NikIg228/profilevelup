#!/usr/bin/env python3
"""
Детальный просмотр структуры файла 12-17.xlsx
"""
import openpyxl
from pathlib import Path

def inspect_excel_detailed(file_path, max_rows=50):
    """Показывает детальную структуру Excel файла"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print(f"\n{'='*80}")
    print(f"Файл: {file_path.name}")
    print(f"{'='*80}")
    
    # Ищем строки с вопросами
    print("\nПоиск строк с вопросами (паттерны: '1', '2', '3', '4', '5'):")
    print("-" * 80)
    
    question_rows = []
    for row_idx, row in enumerate(sheet.iter_rows(max_row=max_rows, values_only=True), start=1):
        if row and row[0]:
            cell_value = str(row[0]).strip()
            # Ищем строки, которые могут быть вопросами
            if any(pattern in cell_value for pattern in ['1️⃣', '1.', '1)', '2️⃣', '2.', '2)', '3️⃣', '3.', '3)', '4️⃣', '4.', '4)', '5️⃣', '5.', '5)']):
                question_rows.append((row_idx, row))
                print(f"\nСтрока {row_idx}:")
                for col_idx, cell_value in enumerate(row[:10], start=1):  # Первые 10 колонок
                    if cell_value:
                        print(f"  Колонка {col_idx}: {repr(str(cell_value)[:100])}")
    
    if not question_rows:
        print("\nВопросы не найдены в первых строках. Показываю все непустые строки:")
        for row_idx, row in enumerate(sheet.iter_rows(max_row=30, values_only=True), start=1):
            if row and any(row):
                print(f"\nСтрока {row_idx}:")
                for col_idx, cell_value in enumerate(row[:5], start=1):
                    if cell_value:
                        print(f"  Колонка {col_idx}: {repr(str(cell_value)[:80])}")

if __name__ == "__main__":
    base_dir = Path(__file__).parent.parent
    excel_path = base_dir / "tests" / "FREE" / "12-17.xlsx"
    inspect_excel_detailed(excel_path, max_rows=100)

