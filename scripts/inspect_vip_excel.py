#!/usr/bin/env python3
"""
Скрипт для просмотра структуры Excel файлов VIP тестов
"""
import openpyxl
from pathlib import Path

def inspect_excel(file_path):
    """Показывает структуру Excel файла"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print(f"\n{'='*80}")
    print(f"Файл: {file_path.name}")
    print(f"{'='*80}")
    print(f"Всего строк: {sheet.max_row}")
    print(f"Всего колонок: {sheet.max_column}")
    print(f"\nПервые 30 строк:")
    print("-" * 80)
    
    for row_idx, row in enumerate(sheet.iter_rows(max_row=30, values_only=True), start=1):
        if row and any(row):
            print(f"\nСтрока {row_idx}:")
            for col_idx, cell_value in enumerate(row[:10], start=1):  # Первые 10 колонок
                if cell_value:
                    print(f"  Колонка {col_idx}: {repr(str(cell_value)[:100])}")

if __name__ == "__main__":
    base_dir = Path(__file__).parent.parent
    tests_dir = base_dir / "tests" / "VIP"
    
    for excel_file in ["12-17.xlsx", "18-20.xlsx", "21+.xlsx"]:
        excel_path = tests_dir / excel_file
        if excel_path.exists():
            inspect_excel(excel_path)

