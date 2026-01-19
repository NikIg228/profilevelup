#!/usr/bin/env python3
"""
Скрипт для просмотра структуры Excel файлов
"""
import openpyxl
from pathlib import Path

def inspect_excel(file_path):
    """Показывает структуру Excel файла"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print(f"\n{'='*60}")
    print(f"Файл: {file_path.name}")
    print(f"{'='*60}")
    print(f"Всего строк: {sheet.max_row}")
    print(f"Всего колонок: {sheet.max_column}")
    print(f"\nПервые 10 строк:")
    print("-" * 60)
    
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
        print(f"Строка {row_idx}:")
        for col_idx, cell_value in enumerate(row, start=1):
            if cell_value:
                print(f"  Колонка {col_idx}: {repr(cell_value)}")
        print()

if __name__ == "__main__":
    base_dir = Path(__file__).parent.parent
    tests_dir = base_dir / "tests" / "FREE"
    
    for excel_file in ["12-17.xlsx", "18-20.xlsx", "21+.xlsx"]:
        excel_path = tests_dir / excel_file
        if excel_path.exists():
            inspect_excel(excel_path)

