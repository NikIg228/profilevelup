#!/usr/bin/env python3
"""
Скрипт для анализа структуры файла text_modules.xlsx
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

def inspect_text_modules(file_path):
    """Анализирует структуру файла text_modules.xlsx"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print(f"📊 Листы в файле: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"📄 Лист: {sheet_name}")
        print(f"   Размер: {sheet.max_row} строк, {sheet.max_column} колонок\n")
        
        # Показываем первые 20 строк для понимания структуры
        print("Первые 20 строк:")
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i > 20:
                break
            row_data = [str(cell) if cell is not None else '' for cell in row[:10]]  # Первые 10 колонок
            print(f"  Строка {i}: {row_data}")
        
        print("\n" + "="*80 + "\n")

if __name__ == "__main__":
    file_path = Path(__file__).parent.parent / "text_modules" / "text_modules.xlsx"
    
    if not file_path.exists():
        print(f"❌ Файл не найден: {file_path}")
        sys.exit(1)
    
    inspect_text_modules(file_path)

