#!/usr/bin/env python3
"""
Поиск вопроса 19 в файле 12-17.xlsx
"""
import openpyxl
from pathlib import Path
import re

def find_question(file_path, question_num):
    """Ищет конкретный вопрос в файле"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print(f"\nПоиск вопроса {question_num} в файле {file_path.name}:")
    print("-" * 80)
    
    found = False
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and row[0]:
            cell_value = str(row[0]).strip()
            # Ищем номер вопроса
            match = re.search(rf'^{question_num}[.)\s]', cell_value)
            if match:
                found = True
                print(f"\nНайден на строке {row_idx}:")
                print(f"  Текст: {cell_value}")
                
                # Показываем следующие строки с вариантами
                for next_row_idx in range(row_idx, min(row_idx + 5, sheet.max_row + 1)):
                    next_row = list(sheet.iter_rows(min_row=next_row_idx, max_row=next_row_idx, values_only=True))[0]
                    if next_row and next_row[0]:
                        option_text = str(next_row[0]).strip()
                        option_letter = str(next_row[1]).strip() if len(next_row) > 1 and next_row[1] else ""
                        print(f"  Строка {next_row_idx}: {option_text[:80]}... | Буква: {option_letter}")
                break
    
    if not found:
        print(f"❌ Вопрос {question_num} не найден")
        print("\nПроверяю все вопросы вокруг 19:")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row and row[0]:
                cell_value = str(row[0]).strip()
                match = re.search(r'^(1[89]|2[0-9])[.)\s]', cell_value)
                if match:
                    print(f"  Строка {row_idx}: {cell_value[:80]}...")

if __name__ == "__main__":
    base_dir = Path(__file__).parent.parent
    excel_path = base_dir / "tests" / "VIP" / "12-17.xlsx"
    find_question(excel_path, 19)

