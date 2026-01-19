#!/usr/bin/env python3
"""
Исправление файла 12-17.json - поиск правильного первого вопроса
"""
import openpyxl
from pathlib import Path

def find_first_question(file_path):
    """Находит правильный первый вопрос"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print("Поиск первого вопроса (должен начинаться с 'Когда перед тобой...'):")
    print("-" * 80)
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and row[0]:
            cell_value = str(row[0]).strip()
            if 'Когда перед тобой' in cell_value:
                print(f"\nНайден вопрос на строке {row_idx}:")
                print(f"  Текст: {cell_value}")
                
                # Показываем следующие строки с вариантами
                for next_row_idx in range(row_idx, min(row_idx + 5, sheet.max_row + 1)):
                    next_row = list(sheet.iter_rows(min_row=next_row_idx, max_row=next_row_idx, values_only=True))[0]
                    if next_row and next_row[0]:
                        print(f"  Строка {next_row_idx}: {str(next_row[0])[:80]}")
                break

if __name__ == "__main__":
    base_dir = Path(__file__).parent.parent
    excel_path = base_dir / "tests" / "FREE" / "12-17.xlsx"
    find_first_question(excel_path)

