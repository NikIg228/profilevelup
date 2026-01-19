#!/usr/bin/env python3
"""
Проверка вопроса 4 во всех файлах
"""
import openpyxl
from pathlib import Path

def check_question_4(file_path):
    """Проверяет вопрос 4 в файле"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print(f"\n{'='*80}")
    print(f"Файл: {file_path.name}")
    print(f"{'='*80}")
    
    found_q4 = False
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and row[0]:
            cell_value = str(row[0]).strip()
            if 'повседневных' in cell_value.lower() or 'повседневной' in cell_value.lower():
                found_q4 = True
                print(f"\nВопрос 4 найден на строке {row_idx}:")
                print(f"  Текст: {cell_value}")
                
                # Показываем следующие строки с вариантами
                for next_row_idx in range(row_idx, min(row_idx + 5, sheet.max_row + 1)):
                    next_row = list(sheet.iter_rows(min_row=next_row_idx, max_row=next_row_idx, values_only=True))[0]
                    if next_row:
                        option_text = str(next_row[0]).strip() if next_row[0] else ""
                        option_letter = str(next_row[1]).strip() if len(next_row) > 1 and next_row[1] else ""
                        if option_text:
                            print(f"  Строка {next_row_idx}: {option_text[:60]}... | Буква: {option_letter}")
                break
    
    if not found_q4:
        print("Вопрос 4 не найден")

if __name__ == "__main__":
    base_dir = Path(__file__).parent.parent
    tests_dir = base_dir / "tests" / "FREE"
    
    for excel_file in ["12-17.xlsx", "18-20.xlsx", "21+.xlsx"]:
        excel_path = tests_dir / excel_file
        if excel_path.exists():
            check_question_4(excel_path)

